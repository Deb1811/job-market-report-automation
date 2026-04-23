import sqlite3
from openpyxl import Workbook
from scripts.transform import transform_data

def create_report():
    print("\n📄 Generating Report...\n")

    data = transform_data()

    wb = Workbook()

    # ---------------- SUMMARY ----------------
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    ws.append(["Total Jobs", data["total_jobs"]])
    ws.append(["Average Salary", data["avg_salary"]])

    # ---------------- TOP COMPANIES ----------------
    ws2 = wb.create_sheet("Top Companies")
    ws2.append(["Company", "Count"])
    for k, v in data["top_companies"].items():
        ws2.append([k, v])

    # ---------------- TOP ROLES ----------------
    ws3 = wb.create_sheet("Top Roles")
    ws3.append(["Role", "Count"])
    for k, v in data["top_roles"].items():
        ws3.append([k, v])

    # ---------------- TOP LOCATIONS ----------------
    ws4 = wb.create_sheet("Top Locations")
    ws4.append(["Location", "Count"])
    for k, v in data["top_locations"].items():
        ws4.append([k, v])

    # ---------------- TOP SKILLS ----------------
    ws5 = wb.create_sheet("Top Skills")
    ws5.append(["Skill", "Count"])
    for k, v in data["top_skills"].items():
        ws5.append([k, v])

    # ---------------- SALARY BY ROLE ----------------
    ws6 = wb.create_sheet("Salary by Role")
    ws6.append(["Role", "Avg Salary"])
    for k, v in data["role_salary"].items():
        ws6.append([k, v])

    # ---------------- JOB TREND ----------------
    ws7 = wb.create_sheet("Job Trends")
    ws7.append(["Month", "Job Count"])
    for k, v in data["trend"].items():
        ws7.append([str(k), v])

    wb.save("reports/job_report.xlsx")

    print("✅ Report generated with 7 sheets!")


if __name__ == "__main__":
    create_report()